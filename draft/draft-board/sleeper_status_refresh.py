#!/usr/bin/env python3
"""
sleeper_status_refresh.py
-------------------------
Morning-of-draft LIVE STATUS refresh for the 2026 ESPN Full-PPR draft board.

What it does
    Pulls the free Sleeper players feed (api.sleeper.app/v1/players/nfl), matches
    every player on the "Draft Board" tab by name, and writes a fresh "Live Status"
    column (injury designation + current team) into a NEW copy of the workbook.
    Cells are shaded red for OUT / IR / PUP / SUS / DOUBTFUL and amber for QUESTIONABLE.

What it does NOT do
    Sleeper's public API does not publish consensus rankings or ADP (its only
    rank-like field, "search_rank", is a popularity signal, not a draft ranking).
    So this refreshes STATUS only — keep the board itself for the actual rankings.
    For rankings/ADP use your ESPN league rankings, FantasyPros, or FFC (full-PPR).

Requirements
    Python 3.8+, internet access, and openpyxl:  pip install openpyxl
Usage
    python sleeper_status_refresh.py "2026_ESPN_FullPPR_Draft_Board.xlsx"
    -> writes  2026_ESPN_FullPPR_Draft_Board_LIVE_<YYYY-MM-DD>.xlsx
"""

import json, re, sys, urllib.request, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

SLEEPER_URL = "https://api.sleeper.app/v1/players/nfl"
RED   = PatternFill("solid", fgColor="FFFCA5A5")
AMBER = PatternFill("solid", fgColor="FFFDE68A")
GREEN = PatternFill("solid", fgColor="FF86EFAC")
OUT_LIKE = {"out", "ir", "pup", "sus", "susp", "doubtful", "dnr", "nfi"}

SUFFIX = re.compile(r"\b(jr|sr|ii|iii|iv|v)\b\.?", re.I)
def norm(name: str) -> str:
    if not name:
        return ""
    n = name.lower().replace("&", "and")
    n = n.replace("’", "'").replace("`", "'")
    n = SUFFIX.sub("", n)
    n = re.sub(r"[^a-z ]", "", n)      # drop punctuation/digits
    return re.sub(r"\s+", " ", n).strip()

def fetch_players() -> dict:
    print("Fetching Sleeper player feed (~5 MB, one call)…")
    req = urllib.request.Request(SLEEPER_URL, headers={"User-Agent": "draft-board-refresh"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.load(r)

def build_index(players: dict) -> dict:
    """normalized full name -> record (prefer offensive/K/DEF fantasy positions)."""
    idx = {}
    for p in players.values():
        full = p.get("full_name") or " ".join(
            x for x in (p.get("first_name"), p.get("last_name")) if x)
        key = norm(full)
        if not key:
            continue
        # keep the most 'fantasy relevant' record if names collide
        prev = idx.get(key)
        if prev is None or (p.get("fantasy_positions") and not prev.get("fantasy_positions")):
            idx[key] = p
    return idx

def status_text(p: dict) -> str:
    inj = (p.get("injury_status") or "").strip()
    team = p.get("team") or "FA"
    st = (p.get("status") or "").strip()          # Active / Inactive / etc.
    parts = []
    if inj:
        parts.append(inj)
    if st and st.lower() not in ("active",) and st.lower() not in (inj.lower(),):
        parts.append(st)
    label = " / ".join(parts) if parts else "Active"
    return f"{label} ({team})"

def severity(p: dict) -> str:
    blob = " ".join(str(p.get(k) or "") for k in ("injury_status", "status")).lower()
    if any(w in blob for w in OUT_LIKE):
        return "red"
    if "questionable" in blob:
        return "amber"
    return "none"

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python sleeper_status_refresh.py <workbook.xlsx>")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"File not found: {src}")

    idx = build_index(fetch_players())
    wb = openpyxl.load_workbook(src)
    ws = wb["Draft Board"]

    # place "Live Status" one column past the current last header
    col = ws.max_column + 1
    h = ws.cell(1, col, "Live Status")
    h.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    h.fill = PatternFill("solid", fgColor="FF1F2937")
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    matched = miss = red = amber = 0
    misses = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 4).value          # column D = Player
        if not name:
            continue
        rec = idx.get(norm(str(name)))
        cell = ws.cell(r, col)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if not rec:
            cell.value = "— no match —"
            miss += 1
            misses.append(str(name))
            continue
        cell.value = status_text(rec)
        sev = severity(rec)
        if sev == "red":
            cell.fill = RED; red += 1
        elif sev == "amber":
            cell.fill = AMBER; amber += 1
        matched += 1

    stamp = datetime.date.today().isoformat()
    out = src.with_name(f"{src.stem}_LIVE_{stamp}.xlsx")
    wb.save(out)
    print(f"\nMatched {matched}  |  flagged {red} OUT-type, {amber} questionable  |  {miss} unmatched")
    if misses:
        print("Unmatched (check spelling / rookies / D-STs):")
        for m in misses:
            print("   -", m)
    print(f"\nSaved -> {out}")
    print("Note: Sleeper provides STATUS, not rankings/ADP — use the board for rankings.")

if __name__ == "__main__":
    main()
